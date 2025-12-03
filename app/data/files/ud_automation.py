#!/usr/bin/env python3
"""
Script d'automatisation pour le Template Tableau de Bord UD CGT
Compatible avec Claude Code pour génération et manipulation de tableaux de bord

Usage:
    python ud_automation.py create --ud 34 --nom "Hérault"
    python ud_automation.py export --fichier TDB_UD66.xlsx --format json
    python ud_automation.py rapport --fichier TDB_UD66.xlsx
    python ud_automation.py sync-dashboard --fichier TDB_UD66.xlsx --api-url https://app.pap-cse.org/api
"""

import pandas as pd
import json
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class UDTableauBord:
    """Classe pour manipuler les tableaux de bord UD"""
    
    FEUILLES_STRUCTURE = {
        'TDB': {
            'colonnes': ['ORGANISATION', 'DATE ET HEURE', 'HEURE', 'RUBRIQUE', 
                        'PRÉSENTATEUR', 'Ordre du jour', 'Référent CEC', 'Collectif confédéral']
        },
        'A CIBLE': {
            'colonnes': ['CIBLE', 'DATE PRESENTATION CE', 'VOTE CE', 'DATE SCRUTIN', 
                        'SUIVI PAP', 'NB SALARIES', 'KBIS ETBS SECONDAIRES', 'Sites',
                        'idcc', 'N° SIRET', 'VOIX CGT', 'NB SYNDIQUES', 'PILOTE', 
                        'OBJET', 'ENJEUX', 'ORGA', 'CONTACT', 'TELEPHONE', 'MAIL']
        },
        'A CIBLE ABSENTE': {
            'colonnes': ['CIBLE', 'DATE PRESENTATION CE', 'VOTE CE', 'DATE SCRUTIN', 
                        'SUIVI PAP', 'NB SALARIES', 'KBIS ETBS SECONDAIRES', 'Sites',
                        'idcc', 'N° SIRET', 'VOIX CGT', 'NB SYNDIQUES', 'PILOTE', 
                        'OBJET', 'ENJEUX', 'ORGA', 'CONTACT', 'TELEPHONE', 'MAIL',
                        'UD', 'ANNEE']
        }
    }
    
    def __init__(self, fichier: str):
        self.fichier = Path(fichier)
        self.wb = None
        
    def creer_nouveau_tableau(self, ud_numero: str, ud_nom: str, fichier_sortie: str):
        """Crée un nouveau tableau de bord vide pour une UD"""
        
        wb = openpyxl.Workbook()
        
        # Créer la feuille TDB
        ws_tdb = wb.active
        ws_tdb.title = f"TDB {ud_numero}"
        self._creer_feuille_tdb(ws_tdb, ud_numero, ud_nom)
        
        # Créer A CIBLE
        ws_cible = wb.create_sheet("A CIBLE")
        self._creer_feuille_structure(ws_cible, self.FEUILLES_STRUCTURE['A CIBLE']['colonnes'])
        
        # Créer A CIBLE ABSENTE
        ws_absente = wb.create_sheet("A CIBLE ABSENTE")
        self._creer_feuille_structure(ws_absente, self.FEUILLES_STRUCTURE['A CIBLE ABSENTE']['colonnes'])
        
        # Note pour les feuilles nationales
        ws_note = wb.create_sheet("NOTE")
        ws_note['A1'] = "Les feuilles AVS et COORDONNEES EVS doivent être copiées depuis le fichier template national"
        ws_note['A1'].font = Font(bold=True, size=12)
        
        wb.save(fichier_sortie)
        print(f"✅ Tableau de bord créé : {fichier_sortie}")
        print(f"📍 UD{ud_numero} - {ud_nom}")
        print(f"⚠️  N'oubliez pas d'ajouter les feuilles AVS et COORDONNEES EVS depuis le template national")
        
    def _creer_feuille_tdb(self, ws, ud_numero: str, ud_nom: str):
        """Crée la structure de la feuille tableau de bord"""
        # Titre
        ws['E2'] = f"TABLEAU DE BORD UD{ud_numero}"
        ws['E2'].font = Font(bold=True, size=16)
        
        # Organisation
        ws['C10'] = "ORGANISATION"
        ws['I10'] = f"ud{ud_numero}"
        ws['C11'] = ud_nom
        
        # En-têtes du tableau
        headers = ['HEURE', 'RUBRIQUE', 'PRÉSENTATEUR']
        for idx, header in enumerate(headers, start=3):
            cell = ws.cell(row=19, column=idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def _creer_feuille_structure(self, ws, colonnes: List[str]):
        """Crée une feuille avec les en-têtes de colonnes"""
        # Titre
        ws['C1'] = ws.title
        ws['C1'].font = Font(bold=True, size=14)
        
        # En-têtes à partir de la ligne 6
        for idx, col in enumerate(colonnes, start=1):
            cell = ws.cell(row=6, column=idx)
            cell.value = col
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def exporter_vers_json(self, feuille: str, fichier_sortie: str):
        """Exporte une feuille vers JSON"""
        df = pd.read_excel(self.fichier, sheet_name=feuille)
        
        # Nettoyer les données
        df = df.dropna(how='all')
        df = df.where(pd.notnull(df), None)
        
        # Convertir les dates en string ISO
        for col in df.select_dtypes(include=['datetime64']).columns:
            df[col] = df[col].astype(str)
        
        data = df.to_dict('records')
        
        with open(fichier_sortie, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Exporté {len(data)} lignes de '{feuille}' vers {fichier_sortie}")
    
    def generer_rapport(self, fichier_sortie: str):
        """Génère un rapport Markdown complet du tableau de bord"""
        
        rapport = []
        rapport.append("# 📊 Rapport Tableau de Bord UD\n")
        rapport.append(f"**Date de génération** : {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
        rapport.append(f"**Fichier source** : {self.fichier.name}\n")
        rapport.append("\n---\n")
        
        # Analyser A CIBLE
        try:
            df_cible = pd.read_excel(self.fichier, sheet_name='A CIBLE')
            df_cible = df_cible.dropna(how='all')
            
            rapport.append("\n## 🎯 Entreprises Cibles (Présence CGT)\n")
            rapport.append(f"**Nombre total** : {len(df_cible)}\n")
            
            # Statistiques
            if 'NB SALARIES' in df_cible.columns:
                total_salaries = df_cible['NB SALARIES'].sum()
                rapport.append(f"**Salariés couverts** : {int(total_salaries):,}\n")
            
            if 'NB SYNDIQUES' in df_cible.columns:
                total_syndiques = df_cible['NB SYNDIQUES'].sum()
                rapport.append(f"**Syndiqués CGT** : {int(total_syndiques)}\n")
                taux = (total_syndiques / total_salaries * 100) if total_salaries > 0 else 0
                rapport.append(f"**Taux de syndicalisation** : {taux:.2f}%\n")
            
            # Prochaines échéances
            if 'DATE SCRUTIN' in df_cible.columns:
                df_cible['DATE SCRUTIN'] = pd.to_datetime(df_cible['DATE SCRUTIN'], errors='coerce')
                aujourd_hui = pd.Timestamp.now()
                dans_90j = aujourd_hui + timedelta(days=90)
                
                prochaines = df_cible[
                    (df_cible['DATE SCRUTIN'] >= aujourd_hui) & 
                    (df_cible['DATE SCRUTIN'] <= dans_90j)
                ]
                
                rapport.append(f"\n### 📅 Élections dans les 90 prochains jours : {len(prochaines)}\n")
                
                if len(prochaines) > 0:
                    rapport.append("\n| Entreprise | Date | SIRET | Pilote |\n")
                    rapport.append("|------------|------|-------|--------|\n")
                    for _, row in prochaines.iterrows():
                        date_str = row['DATE SCRUTIN'].strftime('%d/%m/%Y') if pd.notna(row['DATE SCRUTIN']) else 'N/A'
                        siret = str(row['N° SIRET'])[:14] if pd.notna(row['N° SIRET']) else 'N/A'
                        pilote = row['PILOTE'] if pd.notna(row['PILOTE']) else 'Non assigné'
                        rapport.append(f"| {row['CIBLE']} | {date_str} | {siret} | {pilote} |\n")
            
            # Répartition par pilote
            if 'PILOTE' in df_cible.columns:
                pilotes = df_cible['PILOTE'].value_counts()
                rapport.append(f"\n### 👥 Répartition par pilote\n")
                for pilote, count in pilotes.items():
                    if pd.notna(pilote):
                        rapport.append(f"- **{pilote}** : {count} dossiers\n")
        
        except Exception as e:
            rapport.append(f"\n⚠️ Erreur lors de l'analyse de A CIBLE : {e}\n")
        
        # Analyser A CIBLE ABSENTE
        try:
            df_absente = pd.read_excel(self.fichier, sheet_name='A CIBLE ABSENTE')
            df_absente = df_absente.dropna(how='all')
            
            rapport.append("\n---\n")
            rapport.append("\n## 🚀 Entreprises Cibles (Absence CGT)\n")
            rapport.append(f"**Nombre total** : {len(df_absente)}\n")
            
            if 'NB SALARIES' in df_absente.columns:
                total_salaries_absente = df_absente['NB SALARIES'].sum()
                rapport.append(f"**Potentiel salariés** : {int(total_salaries_absente):,}\n")
            
            # Top 10 par effectif
            if 'NB SALARIES' in df_absente.columns and len(df_absente) > 0:
                top10 = df_absente.nlargest(10, 'NB SALARIES')
                rapport.append("\n### 🔝 Top 10 par effectif\n")
                rapport.append("\n| Entreprise | Effectif | SIRET |\n")
                rapport.append("|------------|----------|-------|\n")
                for _, row in top10.iterrows():
                    siret = str(row['N° SIRET'])[:14] if pd.notna(row['N° SIRET']) else 'N/A'
                    effectif = int(row['NB SALARIES']) if pd.notna(row['NB SALARIES']) else 'N/A'
                    rapport.append(f"| {row['CIBLE']} | {effectif} | {siret} |\n")
        
        except Exception as e:
            rapport.append(f"\n⚠️ Erreur lors de l'analyse de A CIBLE ABSENTE : {e}\n")
        
        # Sauvegarder
        with open(fichier_sortie, 'w', encoding='utf-8') as f:
            f.write(''.join(rapport))
        
        print(f"✅ Rapport généré : {fichier_sortie}")
    
    def synchroniser_dashboard(self, api_url: str, ud_code: str):
        """Prépare les données pour synchronisation avec PAP CSE Dashboard"""
        
        export_data = {
            'ud_code': ud_code,
            'date_export': datetime.now().isoformat(),
            'entreprises': []
        }
        
        try:
            df_cible = pd.read_excel(self.fichier, sheet_name='A CIBLE')
            df_cible = df_cible.dropna(subset=['N° SIRET'])  # Uniquement avec SIRET
            
            for _, row in df_cible.iterrows():
                entreprise = {
                    'siret': str(row['N° SIRET']).replace(' ', ''),
                    'nom_entreprise': row['CIBLE'],
                    'nb_salaries': int(row['NB SALARIES']) if pd.notna(row['NB SALARIES']) else None,
                    'date_election': row['DATE SCRUTIN'].strftime('%Y-%m-%d') if pd.notna(row['DATE SCRUTIN']) else None,
                    'voix_cgt': int(row['VOIX CGT']) if pd.notna(row['VOIX CGT']) else None,
                    'nb_syndiques': int(row['NB SYNDIQUES']) if pd.notna(row['NB SYNDIQUES']) else None,
                    'pilote': row['PILOTE'] if pd.notna(row['PILOTE']) else None,
                    'idcc': str(row['idcc']) if pd.notna(row['idcc']) else None,
                    'source': 'tableau_bord_ud'
                }
                export_data['entreprises'].append(entreprise)
            
            # Sauvegarder en JSON
            fichier_export = f"export_dashboard_{ud_code}_{datetime.now().strftime('%Y%m%d')}.json"
            with open(fichier_export, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            print(f"✅ Export préparé : {fichier_export}")
            print(f"📊 {len(export_data['entreprises'])} entreprises exportées")
            print(f"🔗 À importer dans : {api_url}")
            
            return fichier_export
        
        except Exception as e:
            print(f"❌ Erreur lors de l'export : {e}")
            return None
    
    def valider_donnees(self):
        """Valide la cohérence des données"""
        
        erreurs = []
        
        try:
            df_cible = pd.read_excel(self.fichier, sheet_name='A CIBLE')
            
            # Validation SIRET
            if 'N° SIRET' in df_cible.columns:
                sirets_invalides = df_cible[
                    df_cible['N° SIRET'].notna() & 
                    ~df_cible['N° SIRET'].astype(str).str.match(r'^\d{14}$')
                ]
                if len(sirets_invalides) > 0:
                    erreurs.append(f"❌ {len(sirets_invalides)} SIRET invalides (doivent contenir 14 chiffres)")
            
            # Validation effectifs
            if 'NB SYNDIQUES' in df_cible.columns and 'NB SALARIES' in df_cible.columns:
                incoherence = df_cible[
                    (df_cible['NB SYNDIQUES'].notna()) & 
                    (df_cible['NB SALARIES'].notna()) &
                    (df_cible['NB SYNDIQUES'] > df_cible['NB SALARIES'])
                ]
                if len(incoherence) > 0:
                    erreurs.append(f"❌ {len(incoherence)} entreprises : nb syndiqués > effectif total")
            
            # Validation dates
            if 'DATE SCRUTIN' in df_cible.columns and 'DATE PRESENTATION CE' in df_cible.columns:
                df_cible['DATE SCRUTIN'] = pd.to_datetime(df_cible['DATE SCRUTIN'], errors='coerce')
                df_cible['DATE PRESENTATION CE'] = pd.to_datetime(df_cible['DATE PRESENTATION CE'], errors='coerce')
                
                dates_incoherentes = df_cible[
                    (df_cible['DATE SCRUTIN'].notna()) & 
                    (df_cible['DATE PRESENTATION CE'].notna()) &
                    (df_cible['DATE SCRUTIN'] < df_cible['DATE PRESENTATION CE'])
                ]
                if len(dates_incoherentes) > 0:
                    erreurs.append(f"❌ {len(dates_incoherentes)} dates incohérentes (scrutin avant présentation CE)")
        
        except Exception as e:
            erreurs.append(f"❌ Erreur lors de la validation : {e}")
        
        if len(erreurs) == 0:
            print("✅ Toutes les validations sont passées !")
        else:
            print("\n⚠️  Erreurs détectées :")
            for erreur in erreurs:
                print(f"  {erreur}")
        
        return len(erreurs) == 0


def main():
    parser = argparse.ArgumentParser(description='Automatisation Tableau de Bord UD CGT')
    subparsers = parser.add_subparsers(dest='command', help='Commandes disponibles')
    
    # Commande create
    create_parser = subparsers.add_parser('create', help='Créer un nouveau tableau de bord')
    create_parser.add_argument('--ud', required=True, help='Numéro du département (ex: 34)')
    create_parser.add_argument('--nom', required=True, help='Nom du département (ex: Hérault)')
    create_parser.add_argument('--output', default=None, help='Nom du fichier de sortie')
    
    # Commande export
    export_parser = subparsers.add_parser('export', help='Exporter vers JSON')
    export_parser.add_argument('--fichier', required=True, help='Fichier Excel source')
    export_parser.add_argument('--feuille', default='A CIBLE', help='Feuille à exporter')
    export_parser.add_argument('--output', default=None, help='Fichier JSON de sortie')
    
    # Commande rapport
    rapport_parser = subparsers.add_parser('rapport', help='Générer un rapport')
    rapport_parser.add_argument('--fichier', required=True, help='Fichier Excel source')
    rapport_parser.add_argument('--output', default=None, help='Fichier Markdown de sortie')
    
    # Commande sync-dashboard
    sync_parser = subparsers.add_parser('sync-dashboard', help='Synchroniser avec PAP CSE Dashboard')
    sync_parser.add_argument('--fichier', required=True, help='Fichier Excel source')
    sync_parser.add_argument('--ud-code', required=True, help='Code UD (ex: ud66)')
    sync_parser.add_argument('--api-url', default='https://app.pap-cse.org/api', help='URL de l\'API')
    
    # Commande validate
    validate_parser = subparsers.add_parser('validate', help='Valider les données')
    validate_parser.add_argument('--fichier', required=True, help='Fichier Excel source')
    
    args = parser.parse_args()
    
    if args.command == 'create':
        ud = UDTableauBord('dummy.xlsx')
        output = args.output or f"TABLEAU_de_BORD_UD{args.ud}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        ud.creer_nouveau_tableau(args.ud, args.nom, output)
    
    elif args.command == 'export':
        ud = UDTableauBord(args.fichier)
        output = args.output or f"{args.feuille.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.json"
        ud.exporter_vers_json(args.feuille, output)
    
    elif args.command == 'rapport':
        ud = UDTableauBord(args.fichier)
        output = args.output or f"rapport_{Path(args.fichier).stem}_{datetime.now().strftime('%Y%m%d')}.md"
        ud.generer_rapport(output)
    
    elif args.command == 'sync-dashboard':
        ud = UDTableauBord(args.fichier)
        ud.synchroniser_dashboard(args.api_url, args.ud_code)
    
    elif args.command == 'validate':
        ud = UDTableauBord(args.fichier)
        ud.valider_donnees()
    
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
