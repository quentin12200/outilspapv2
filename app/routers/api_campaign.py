"""
API Router pour la gestion de campagnes PAP.

Ce module permet d'analyser des PAP en masse, de les classer par priorité
et de générer les contenus d'emails pour les UD.
"""

import logging
import os
import uuid
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Request
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..db import get_session
from ..services.document_extractor import DocumentExtractor, DocumentExtractorError
from ..services.pap_campaign_service import PAPCampaignService
from ..services.pap_enrichment_service import PAPEnrichmentService
from ..services.pappers_api import PappersAPI
from ..services.idcc_enrichment import get_idcc_enrichment_service
from ..audit import log_admin_action
from ..user_auth import require_admin_user
from ..models import User

logger = logging.getLogger(__name__)

# Répertoire de stockage des PDFs
PAP_UPLOADS_DIR = Path("app/static/pap_uploads")
PAP_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

router = APIRouter(prefix="/api/campaign", tags=["Campagnes PAP"])


class CampaignAnalysisResult(BaseModel):
    """Résultat d'analyse d'une campagne PAP."""
    total: int
    stats: Dict[str, Any]
    enjeux: List[Dict[str, Any]]
    standard: List[Dict[str, Any]]


class EmailGenerationRequest(BaseModel):
    """Requête de génération d'email."""
    pap_data: Dict[str, Any]
    is_priority: bool = False


@router.post("/analyze-batch")
async def analyze_pap_batch(
    request: Request,
    files: List[UploadFile] = File(..., description="Liste de PDFs de PAP à analyser"),
    db: Session = Depends(get_session),
    current_user: User = Depends(require_admin_user)
):
    """
    Analyse un lot de PAP et les classe par priorité (enjeux vs standard).

    **Fonctionnalités:**
    - Extraction automatique des données de chaque PAP (GPT-4 Vision)
    - Enrichissement avec APIs Pappers/Sirene si données manquantes
    - Classification automatique (PAP à enjeux ou standard)
    - Identification de l'UD responsable

    **Critères PAP à enjeux:**
    - Effectif ≥ 1000 salariés
    - OU inscrits importants par rapport à la moyenne du département

    **Exemple:**
    ```bash
    curl -X POST "http://localhost:8000/api/campaign/analyze-batch" \\
      -F "files=@pap1.pdf" \\
      -F "files=@pap2.pdf" \\
      -F "files=@pap3.pdf"
    ```
    """
    if not files:
        raise HTTPException(status_code=400, detail="Aucun fichier fourni")

    extracted_paps = []
    extraction_errors = []

    enrichment_service = PAPEnrichmentService(db)

    logger.info(f"🚀 Début d'analyse de campagne PAP - {len(files)} fichier(s)")

    # Étape 1 : Extraction et enrichissement complet de chaque PAP
    for i, file in enumerate(files, 1):
        logger.info(f"📄 Traitement du fichier {i}/{len(files)}: {file.filename}")

        try:
            # Vérifier le type de fichier
            if file.content_type not in [
                "image/jpeg", "image/jpg", "image/png", "image/webp",
                "application/pdf"
            ]:
                extraction_errors.append({
                    'filename': file.filename,
                    'error': f"Type de fichier non supporté: {file.content_type}"
                })
                continue

            # Lire le fichier
            file_data = await file.read()
            is_pdf = file.content_type == "application/pdf"

            # Extraire et enrichir les informations (GPT-4 + Pappers + Base PV)
            extracted_data = await enrichment_service.enrich_pap_from_pdf(
                file_data,
                file.filename,
                is_pdf=is_pdf
            )

            # Stocker le PDF avec nom basé sur SIRET
            pdf_url = None
            pdf_filename = None
            if is_pdf:
                try:
                    # Utiliser le SIRET pour nommer le fichier si disponible
                    siret = extracted_data.get('siret')
                    if siret and siret.replace(' ', '').isdigit():
                        # Format: SIRET_DATE.pdf (ex: 12345678901234_20250130.pdf)
                        date_str = datetime.now().strftime("%Y%m%d")
                        pdf_filename = f"{siret.replace(' ', '')}_{date_str}.pdf"
                    else:
                        # Fallback sur UUID si pas de SIRET valide
                        pdf_filename = f"{uuid.uuid4()}.pdf"

                    pdf_path = PAP_UPLOADS_DIR / pdf_filename

                    # Sauvegarder le PDF
                    with open(pdf_path, 'wb') as f:
                        f.write(file_data)

                    # Générer l'URL publique
                    pdf_url = f"/static/pap_uploads/{pdf_filename}"
                    logger.info(f"✅ PDF stocké : {pdf_filename}")
                except Exception as e:
                    logger.error(f"⚠️ Erreur stockage PDF {file.filename}: {str(e)}")

            # Ajouter les métadonnées du PDF
            extracted_data['pdf_url'] = pdf_url
            extracted_data['pdf_filename'] = pdf_filename
            extracted_data['original_filename'] = file.filename

            extracted_paps.append(extracted_data)
            logger.info(f"✅ Enrichissement complet - SIRET: {extracted_data.get('siret', 'N/A')}")

        except DocumentExtractorError as e:
            logger.error(f"❌ Erreur extraction {file.filename}: {str(e)}")
            extraction_errors.append({
                'filename': file.filename,
                'error': str(e)
            })
        except Exception as e:
            logger.error(f"❌ Erreur inattendue {file.filename}: {str(e)}")
            extraction_errors.append({
                'filename': file.filename,
                'error': f"Erreur inattendue: {str(e)}"
            })

    if not extracted_paps:
        raise HTTPException(
            status_code=422,
            detail=f"Aucun PAP n'a pu être extrait. Erreurs: {extraction_errors}"
        )

    # Étape 2 : Analyse et classification des PAP (déjà fait dans enrichment_service)
    logger.info(f"🔍 Classification finale de {len(extracted_paps)} PAP(s)")
    campaign_service = PAPCampaignService(db)
    analysis_result = campaign_service.analyze_batch(extracted_paps)

    # Ajouter les erreurs d'extraction dans les stats
    analysis_result['extraction_errors'] = extraction_errors
    analysis_result['extraction_success_rate'] = (
        len(extracted_paps) / len(files) * 100 if files else 0
    )

    # Log de l'action
    log_admin_action(
        request=request,
        api_key=None,
        action="analyze_pap_campaign",
        resource_type="campaign",
        success=True,
        resource_id=f"batch_{len(files)}",
        request_params={
            'total_files': len(files),
            'successful_extractions': len(extracted_paps),
            'failed_extractions': len(extraction_errors)
        },
        response_summary={
            'enjeux': analysis_result['stats']['count_enjeux'],
            'standard': analysis_result['stats']['count_standard']
        }
    )

    logger.info(
        f"✅ Analyse terminée - "
        f"{analysis_result['stats']['count_enjeux']} enjeux, "
        f"{analysis_result['stats']['count_standard']} standard"
    )

    return analysis_result


@router.post("/generate-email")
async def generate_email_content(
    request: Request,
    email_request: EmailGenerationRequest
):
    """
    Génère le contenu d'un email pour un PAP spécifique.

    Le contenu est différencié selon que le PAP est prioritaire ou non.

    **Returns:**
    ```json
    {
        "subject": "Sujet de l'email",
        "body": "Corps de l'email pré-formaté",
        "recipient": "ud75@cgt.fr",
        "mailto_link": "mailto:ud75@cgt.fr?subject=..."
    }
    ```
    """
    try:
        pap_data = email_request.pap_data
        is_priority = email_request.is_priority

        email_content = PAPCampaignService.generate_email_content(pap_data, is_priority)

        # Générer le lien mailto:
        import urllib.parse
        mailto_link = (
            f"mailto:{email_content['recipient']}"
            f"?subject={urllib.parse.quote(email_content['subject'])}"
            f"&body={urllib.parse.quote(email_content['body'])}"
        )

        return {
            **email_content,
            'mailto_link': mailto_link
        }

    except Exception as e:
        logger.error(f"Erreur génération email: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/pappers/{siret}")
async def get_pappers_data(
    siret: str,
    request: Request,
    current_user: User = Depends(require_admin_user)
):
    """
    Interroge l'API Pappers pour récupérer les informations complètes d'une entreprise.

    Args:
        siret: Numéro SIRET de l'entreprise (14 chiffres)

    Returns:
        Informations complètes de l'entreprise depuis Pappers
    """
    try:
        # Nettoyer le SIRET
        siret_clean = ''.join(c for c in siret if c.isdigit())

        if len(siret_clean) != 14:
            raise HTTPException(status_code=400, detail="SIRET invalide (doit contenir 14 chiffres)")

        # Interroger Pappers directement avec httpx pour obtenir toutes les données brutes
        import httpx

        pappers = PappersAPI()
        if not pappers.api_key:
            raise HTTPException(status_code=503, detail="Clé API Pappers non configurée")

        url = f"https://api.pappers.fr/v2/entreprise"
        params = {
            "api_token": pappers.api_key,
            "siret": siret_clean
        }

        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(url, params=params)

            if response.status_code == 404:
                raise HTTPException(status_code=404, detail="Entreprise non trouvée dans Pappers")

            if response.status_code != 200:
                logger.error(f"Erreur API Pappers ({response.status_code}): {response.text[:200]}")
                raise HTTPException(status_code=502, detail=f"Erreur API Pappers (code {response.status_code})")

            pappers_data = response.json()

        logger.info(f"✅ Données Pappers récupérées pour {siret_clean}")

        return {
            "success": True,
            "siret": siret_clean,
            "data": pappers_data
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de l'interrogation Pappers: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")


@router.post("/export-excel")
async def export_paps_to_excel(
    request: Request,
    paps_data: Dict[str, Any]
):
    """
    Exporte les données PAP extraites vers un fichier Excel.

    Args:
        paps_data: Dictionnaire contenant la liste des PAP à exporter

    Returns:
        Fichier Excel téléchargeable
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    try:
        paps = paps_data.get('paps', [])
        if not paps:
            raise HTTPException(status_code=400, detail="Aucune donnée PAP à exporter")

        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "PAP Extraits"

        # Définir les styles
        header_fill = PatternFill(start_color="E31F26", end_color="E31F26", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        enjeux_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")

        # En-têtes dans l'ordre spécifié par l'utilisateur
        headers = [
            # Colonnes principales (ordre imposé)
            "SIRET",
            "Nom Entreprise",
            "CP",
            "Ville",
            "date d'arrivée",
            "UD",
            "FD",
            "Siège social",
            "IDCC",
            "Nombre de salariés",
            "Commentaires",
            "Date de saisie",
            "ENJEUX",
            # Colonnes supplémentaires utiles
            "Adresse",
            "Inscrits",
            "Date Élection",
            "Convention Collective",
            "Type Scrutin",
            "Département",
            "Historique PV",
            "Présence CGT",
            "Lien PDF",
            "Fichier Original",
            "Sources Enrichissement"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Remplir les données
        for row_num, pap in enumerate(paps, 2):
            historique_pv = pap.get('historique_pv', {})
            has_pv = 'Oui' if historique_pv.get('found') else 'Non'

            # Vérifier présence CGT
            presence_cgt = ''
            if historique_pv.get('found'):
                cgt_c3 = historique_pv.get('presence_cgt_c3', False)
                cgt_c4 = historique_pv.get('presence_cgt_c4', False)
                if cgt_c3 and cgt_c4:
                    presence_cgt = 'C3 + C4'
                elif cgt_c3:
                    presence_cgt = 'C3'
                elif cgt_c4:
                    presence_cgt = 'C4'
                else:
                    presence_cgt = 'Non'

            enrichment_sources = ', '.join(pap.get('enrichment_sources', []))

            # Siège social
            est_siege = 'Oui' if pap.get('est_siege') else 'Non'

            # Date de saisie (timestamp actuel)
            date_saisie = datetime.now().strftime("%d/%m/%Y %H:%M")

            # Catégorie
            categorie = 'ENJEUX' if pap.get('category') == 'enjeux' else 'Standard'

            row_data = [
                # Colonnes principales
                pap.get('siret', ''),
                pap.get('raison_sociale', ''),
                pap.get('code_postal', ''),
                pap.get('ville', ''),
                pap.get('date_invitation', ''),
                pap.get('ud', ''),
                pap.get('fd', ''),
                est_siege,
                pap.get('idcc', ''),
                pap.get('effectif', ''),
                pap.get('notes', ''),
                date_saisie,
                categorie,
                # Colonnes supplémentaires
                pap.get('adresse', ''),
                pap.get('inscrits', ''),
                pap.get('date_election', ''),
                pap.get('convention_collective', ''),
                pap.get('type_scrutin', ''),
                pap.get('departement', ''),
                has_pv,
                presence_cgt,
                pap.get('pdf_url', ''),
                pap.get('original_filename', ''),
                enrichment_sources
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Colorer les PAP à enjeux
                if pap.get('category') == 'enjeux':
                    cell.fill = enjeux_fill

        # Ajuster les largeurs de colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Générer le nom de fichier avec timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PAP_extraits_{timestamp}.xlsx"

        logger.info(f"✅ Export Excel généré: {len(paps)} PAP - {filename}")

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Erreur lors de l'export Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export: {str(e)}")


@router.post("/import-excel")
async def import_excel_and_generate_emails(
    request: Request,
    file: UploadFile = File(..., description="Fichier Excel complété à importer"),
    db: Session = Depends(get_session),
    current_user: User = Depends(require_admin_user)
):
    """
    Importe un fichier Excel complété et génère tous les emails pour envoi manuel.

    Args:
        file: Fichier Excel avec données PAP complétées

    Returns:
        Liste de tous les emails générés avec leur contenu complet
    """
    from openpyxl import load_workbook
    from io import BytesIO

    try:
        # Vérifier le type de fichier
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Le fichier doit être au format Excel (.xlsx ou .xls)")

        # Lire le fichier Excel
        file_data = await file.read()
        excel_buffer = BytesIO(file_data)

        try:
            wb = load_workbook(excel_buffer, data_only=True)
            ws = wb.active
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Impossible de lire le fichier Excel: {str(e)}")

        # Lire les en-têtes (ligne 1)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # Trouver les indices des colonnes importantes
        try:
            idx_siret = headers.index("SIRET")
            idx_nom = headers.index("Nom Entreprise")
            idx_cp = headers.index("CP")
            idx_ville = headers.index("Ville")
            idx_date_arrivee = headers.index("date d'arrivée")
            idx_ud = headers.index("UD")
            idx_fd = headers.index("FD")
            idx_siege = headers.index("Siège social")
            idx_idcc = headers.index("IDCC")
            idx_effectif = headers.index("Nombre de salariés")
            idx_commentaires = headers.index("Commentaires")
            idx_enjeux = headers.index("ENJEUX")
            idx_date_invitation = headers.index("Date d'invitation") if "Date d'invitation" in headers else None
            idx_inscrits = headers.index("Inscrits") if "Inscrits" in headers else None
            idx_date_election = headers.index("Date Élection") if "Date Élection" in headers else None
            idx_pdf_url = headers.index("Lien PDF") if "Lien PDF" in headers else None
        except ValueError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Colonne manquante dans le fichier Excel: {str(e)}"
            )

        # Parser les lignes et générer les emails
        paps_with_emails = []
        enrichment_service = PAPEnrichmentService(db)
        campaign_service = PAPCampaignService(db)
        idcc_service = get_idcc_enrichment_service()
        fd_enriched_count = 0  # Compteur pour les FD enrichies automatiquement

        logger.info(f"📥 Import Excel - Traitement de {ws.max_row - 1} lignes")

        for row_num in range(2, ws.max_row + 1):  # Skip header
            row = ws[row_num]

            # Extraire les données de la ligne
            siret = str(row[idx_siret].value or '').strip()
            nom_entreprise = str(row[idx_nom].value or '').strip()
            code_postal = str(row[idx_cp].value or '').strip()
            ville = str(row[idx_ville].value or '').strip()
            date_arrivee = str(row[idx_date_arrivee].value or '').strip()
            ud = str(row[idx_ud].value or '').strip()
            fd = str(row[idx_fd].value or '').strip()
            est_siege = str(row[idx_siege].value or '').strip()
            idcc = str(row[idx_idcc].value or '').strip()
            effectif_str = str(row[idx_effectif].value or '').strip()
            commentaires = str(row[idx_commentaires].value or '').strip()
            enjeux_str = str(row[idx_enjeux].value or '').strip()

            # Enrichissement automatique de la FD à partir de l'IDCC si FD manquante
            if idcc and not fd:
                fd_enriched = idcc_service.enrich_fd(idcc, fd, db)
                if fd_enriched and fd_enriched != fd:
                    fd = fd_enriched
                    fd_enriched_count += 1
                    logger.info(f"  ✨ Ligne {row_num}: FD enrichie automatiquement '{fd}' depuis IDCC {idcc}")

            # Colonnes optionnelles
            date_invitation = str(row[idx_date_invitation].value or '').strip() if idx_date_invitation else ''
            inscrits_str = str(row[idx_inscrits].value or '').strip() if idx_inscrits else ''
            date_election = str(row[idx_date_election].value or '').strip() if idx_date_election else ''
            pdf_url = str(row[idx_pdf_url].value or '').strip() if idx_pdf_url else ''

            # Skip les lignes vides
            if not siret or not nom_entreprise:
                continue

            # Nettoyer et valider le SIRET
            siret_clean = ''.join(c for c in siret if c.isdigit())
            if len(siret_clean) != 14:
                logger.warning(f"⚠️ Ligne {row_num}: SIRET invalide '{siret}' - ignorée")
                continue

            # Convertir effectif et inscrits en nombres
            try:
                effectif = int(effectif_str) if effectif_str and effectif_str.isdigit() else None
            except:
                effectif = None

            try:
                inscrits = int(inscrits_str) if inscrits_str and inscrits_str.isdigit() else None
            except:
                inscrits = None

            # Déterminer si c'est un PAP à enjeux
            is_enjeux = enjeux_str.upper() in ['ENJEUX', 'OUI', 'TRUE', '1']

            # Récupérer l'historique PV pour ce SIRET
            siret_summary = enrichment_service._get_siret_summary(siret_clean)
            historique_pv = {}

            if siret_summary:
                historique_pv = enrichment_service._format_siret_summary(siret_summary)
                # Récupérer aussi les événements et invitations
                pv_events = enrichment_service._get_pv_events(siret_clean)
                old_invitations = enrichment_service._get_old_invitations(siret_clean)
                if pv_events:
                    historique_pv['evenements'] = pv_events
                if old_invitations:
                    historique_pv['anciennes_invitations'] = old_invitations
            else:
                historique_pv = {
                    'found': False,
                    'message': 'Aucun historique PV trouvé dans la base'
                }

            # Construire l'objet PAP
            pap_data = {
                'siret': siret_clean,
                'raison_sociale': nom_entreprise,
                'code_postal': code_postal,
                'ville': ville,
                'date_invitation': date_invitation if date_invitation else date_arrivee,  # Nouvelle colonne ou fallback
                'date_arrivee': date_arrivee,
                'ud': ud,
                'fd': fd,
                'est_siege': est_siege.upper() in ['OUI', 'TRUE', '1'],
                'idcc': idcc,
                'effectif': effectif,
                'inscrits': inscrits,
                'date_election': date_election,
                'notes': commentaires,
                'category': 'enjeux' if is_enjeux else 'standard',
                'is_priority': is_enjeux,
                'pdf_url': pdf_url,
                'historique_pv': historique_pv,
                'departement': code_postal[:2] if code_postal and len(code_postal) >= 2 else None
            }

            # Générer l'email pour ce PAP
            email_content = campaign_service.generate_email_content(pap_data, is_priority=is_enjeux)

            paps_with_emails.append({
                'pap': pap_data,
                'email': email_content,
                'row_number': row_num
            })

            logger.info(f"✅ Ligne {row_num}: Email généré pour {nom_entreprise} ({siret_clean})")

        if not paps_with_emails:
            raise HTTPException(
                status_code=400,
                detail="Aucun PAP valide trouvé dans le fichier Excel"
            )

        # Log de l'action
        log_admin_action(
            request=request,
            api_key=None,
            action="import_excel_pap",
            resource_type="campaign",
            success=True,
            resource_id=f"import_{len(paps_with_emails)}",
            request_params={
                'filename': file.filename,
                'total_rows': ws.max_row - 1,
                'emails_generated': len(paps_with_emails)
            }
        )

        logger.info(f"✅ Import Excel terminé - {len(paps_with_emails)} emails générés")
        if fd_enriched_count > 0:
            logger.info(f"   ✨ {fd_enriched_count} FD enrichies automatiquement depuis IDCC")

        return {
            'success': True,
            'total': len(paps_with_emails),
            'emails': paps_with_emails,
            'stats': {
                'enjeux': sum(1 for p in paps_with_emails if p['pap']['category'] == 'enjeux'),
                'standard': sum(1 for p in paps_with_emails if p['pap']['category'] == 'standard'),
                'with_cgt_history': sum(1 for p in paps_with_emails if p['pap']['historique_pv'].get('found')),
                'new_implantations': sum(1 for p in paps_with_emails if not p['pap']['historique_pv'].get('found')),
                'fd_auto_enriched': fd_enriched_count
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de l'import Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'import: {str(e)}")


@router.get("/health")
async def health_check():
    """Vérifie que le service de campagne est opérationnel."""
    return {
        "status": "operational",
        "message": "Service de campagne PAP prêt"
    }
