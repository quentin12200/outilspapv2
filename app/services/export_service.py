
from datetime import datetime, date
from io import BytesIO
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models import PVEvent
from app.services.calcul_elus_cse import calculer_elus_cse_complet, calculer_nombre_elus_cse, repartir_sieges_quotient_seul

# Constantes
PV_ORGANISATION_FIELDS = [
    ("cgt_voix", "CGT"),
    ("cfdt_voix", "CFDT"),
    ("fo_voix", "FO"),
    ("cftc_voix", "CFTC"),
    ("cgc_voix", "CFE-CGC"),
    ("unsa_voix", "UNSA"),
    ("sud_voix", "Solidaires"),
    ("autre_voix", "Autre"),
]

def _to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(" ", "").replace(",", ".")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None

def _parse_date(value: str | None) -> date | None:
    if not value:
        return None

    cleaned = value.strip()
    if not cleaned:
        return None

    formats = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
    )

    for fmt in formats:
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue

    # Tentative ISO 8601 générique (permet 2025-03-01T00:00:00)
    try:
        return datetime.fromisoformat(cleaned).date()
    except ValueError:
        return None

def _format_int_fr(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            rounded = int(round(float(value)))
        except (TypeError, ValueError):
            return None
    else:
        try:
            rounded = int(round(float(str(value).replace(",", "."))))
        except (TypeError, ValueError):
            return None

    return f"{rounded:,}".replace(",", "\u202f")

def _format_percent_fr(value: float | None, decimals: int = 1) -> str | None:
    if value is None:
        return None
    formatted = f"{value:.{decimals}f}".replace(".", ",")
    return f"{formatted} %"

def generate_calendrier_excel(
    db: Session,
    filters: Dict[str, Any]
) -> BytesIO:
    """
    Génère le fichier Excel pour l'export du calendrier +1000.
    
    Args:
        db: Session SQLAlchemy
        filters: Dictionnaire des filtres (min_effectif, q, cycle, etc.)
        
    Returns:
        BytesIO: Le buffer contenant le fichier Excel
    """
    today = date.today()
    
    # Extraction des filtres
    min_effectif = filters.get("min_effectif", 1000)
    q = filters.get("q", "")
    cycle = filters.get("cycle", "")
    institution = filters.get("institution", "")
    fd = filters.get("fd", "")
    idcc = filters.get("idcc", "")
    ud = filters.get("ud", "")
    region = filters.get("region", "")
    year = filters.get("year", "")

    stmt = (
        db.query(
            PVEvent.siret,
            PVEvent.raison_sociale,
            PVEvent.ud,
            PVEvent.region,
            PVEvent.effectif_siret,
            PVEvent.inscrits,
            PVEvent.cycle,
            PVEvent.date_prochain_scrutin,
            PVEvent.date_pv,
            PVEvent.institution,
            PVEvent.fd,
            PVEvent.idcc,
            PVEvent.sve,
            PVEvent.tx_participation_pv,
            PVEvent.votants,
            PVEvent.nb_college_siret,
            PVEvent.cgt_voix,
            PVEvent.cfdt_voix,
            PVEvent.fo_voix,
            PVEvent.cftc_voix,
            PVEvent.cgc_voix,
            PVEvent.unsa_voix,
            PVEvent.sud_voix,
            PVEvent.autre_voix,
        )
        .filter(PVEvent.date_prochain_scrutin.isnot(None))
    )

    search_term = q.strip().lower() if q else ""
    cycle_filter = cycle.strip() if cycle else ""
    institution_filter = institution.strip() if institution else ""
    fd_filter = fd.strip() if fd else ""
    idcc_filter = idcc.strip() if idcc else ""
    ud_filter = ud.strip() if ud else ""
    region_filter = region.strip() if region else ""
    year_filter = year.strip() if year else ""

    # ÉTAPE 1 : Calculer pour CHAQUE collège/PV (ne pas dédupliquer encore)
    per_siret: dict[str, dict[str, Any]] = {}
    for row in stmt:
        parsed_date = _parse_date(row.date_prochain_scrutin)
        if not parsed_date or parsed_date < today:
            continue

        # Pour le filtre et l'affichage : utiliser effectif_siret ou inscrits
        effectif_siret_value = _to_number(row.effectif_siret)
        effectif_college = _to_number(row.inscrits)  # Effectif du collège

        filter_effectif = effectif_siret_value if effectif_siret_value is not None else effectif_college

        if min_effectif and (filter_effectif is None or filter_effectif < int(min_effectif)):
            continue

        if cycle_filter and (row.cycle or "") != cycle_filter:
            continue
        if institution_filter and (row.institution or "") != institution_filter:
            continue
        if fd_filter and (row.fd or "") != fd_filter:
            continue
        if idcc_filter and (str(row.idcc or "")) != idcc_filter:
            continue
        if ud_filter and (row.ud or "") != ud_filter:
            continue
        if region_filter and (row.region or "") != region_filter:
            continue
        if year_filter and str(parsed_date.year) != year_filter:
            continue

        if search_term:
            siret_value = str(row.siret or "")
            raison = (row.raison_sociale or "").lower()
            if search_term not in siret_value.lower() and search_term not in raison:
                continue

        sve_value = _to_number(getattr(row, "sve", None))
        participation_value = _to_number(getattr(row, "tx_participation_pv", None))

        # Si tx_participation_pv est vide, calculer à partir de votants/inscrits
        if participation_value is None:
            votants_value = _to_number(getattr(row, "votants", None))
            inscrits_value = _to_number(row.inscrits)
            if votants_value is not None and inscrits_value is not None and inscrits_value > 0:
                participation_value = (votants_value / inscrits_value) * 100

        nb_college_value = _to_number(getattr(row, "nb_college_siret", None))

        # Calculer les voix par organisation pour ce collège
        voix_par_orga = {}
        for attr, label in PV_ORGANISATION_FIELDS:
            votes_value = _to_number(getattr(row, attr, None))
            if votes_value and votes_value > 0:
                voix_par_orga[label] = votes_value

        # Calculer les élus CSE pour ce collège (uniquement C4, plafonné à 35 sièges pour 10 000+)
        # IMPORTANT: Utiliser l'effectif DU COLLÈGE (inscrits), PAS l'effectif total entreprise (effectif_siret)
        elus_par_orga = {}
        nb_sieges_cse = None

        if row.cycle == "C4" and effectif_college and effectif_college > 0 and voix_par_orga:
            calcul_elus = calculer_elus_cse_complet(
                int(effectif_college),  # Effectif du collège (inscrits) - JAMAIS effectif_siret !
                {label: int(v) for label, v in voix_par_orga.items()}
            )
            nb_sieges_cse = calcul_elus["nb_sieges_total"]
            elus_par_orga = calcul_elus["elus_par_orga"]

        # Créer une clé unique par collège pour garder tous les collèges
        college_key = f"{row.siret or 'pv'}_{row.cycle or 'na'}_{id(row)}"

        # Récupérer aussi votants et inscrits pour l'agrégation de la participation
        votants_college = _to_number(getattr(row, "votants", None)) or 0
        inscrits_college = _to_number(row.inscrits) or 0

        college_data = {
            "siret": row.siret,
            "raison_sociale": row.raison_sociale,
            "ud": row.ud,
            "region": row.region,
            "effectif_siret": effectif_siret_value,
            "inscrits": inscrits_college,
            "votants": votants_college,
            "cycle": row.cycle,
            "date_prochain_scrutin": parsed_date,
            "date_pv": _parse_date(row.date_pv),
            "institution": row.institution,
            "fd": row.fd,
            "idcc": row.idcc,
            "sve": sve_value,
            "participation": participation_value,
            "nb_college_siret": nb_college_value,
            "voix_par_orga": voix_par_orga,
            "elus_par_orga": elus_par_orga,
            "nb_sieges_cse": nb_sieges_cse,
            # Pour le calcul agrégé
            "effectif": filter_effectif,
        }

        if row.siret not in per_siret:
            per_siret[row.siret] = {}
        per_siret[row.siret][college_key] = college_data

    # ÉTAPE 2 : Agréger par SIRET
    siret_aggregated = {}

    for siret, colleges in per_siret.items():
        # Prendre les infos générales du premier collège trouvé
        first_college = list(colleges.values())[0]
        
        siret_aggregated[siret] = {
            "siret": first_college["siret"],
            "raison_sociale": first_college["raison_sociale"],
            "ud": first_college["ud"],
            "region": first_college["region"],
            "effectif_siret": first_college["effectif_siret"] or 0,  # Effectif entreprise
            "effectif": first_college["effectif"],  # Effectif utilisé pour le filtre
            "cycle": first_college["cycle"],
            "date": first_college["date_prochain_scrutin"],
            "date_display": first_college["date_prochain_scrutin"].strftime("%d/%m/%Y"),
            "date_pv": first_college["date_pv"],
            "institution": first_college["institution"],
            "fd": first_college["fd"],
            "idcc": first_college["idcc"],
            "nb_college": first_college["nb_college_siret"],
            # Totaux à calculer
            "sve": 0,
            "inscrits": 0,
            "votants": 0,
            "voix_par_orga": defaultdict(int),
            "elus_par_orga": defaultdict(int),
            "nb_sieges_cse": 0,
            "colleges_details": []
        }

        # Agréger les données de tous les collèges de ce SIRET
        for college_data in colleges.values():
            # Vérifier le quorum du collège AVANT d'agréger ses votes
            # Le quorum est atteint si : SVE >= (inscrits / 2) + 1
            # Si le quorum n'est pas atteint, ce collège n'a pas d'élus et ses voix ne comptent pas
            college_inscrits = college_data["inscrits"]
            college_sve = college_data["sve"] or 0
            quorum_atteint = False

            if college_inscrits > 0:
                quorum_requis = (college_inscrits / 2) + 1
                quorum_atteint = college_sve >= quorum_requis

            # Additionner les valeurs de ce collège aux totaux du SIRET
            # UNIQUEMENT si le quorum est atteint
            if quorum_atteint:
                siret_aggregated[siret]["sve"] += college_sve
                siret_aggregated[siret]["votants"] += college_data["votants"]
                siret_aggregated[siret]["inscrits"] += college_data["inscrits"]

                for orga, voix in college_data["voix_par_orga"].items():
                    siret_aggregated[siret]["voix_par_orga"][orga] += voix

            siret_aggregated[siret]["colleges_details"].append({
                "effectif": college_data["effectif"],
                "cycle": college_data["cycle"],
                "sve": college_data["sve"],
                "nb_sieges": college_data["nb_sieges_cse"],
                "voix_par_orga": dict(college_data["voix_par_orga"]),
                "elus_par_orga": dict(college_data["elus_par_orga"]),
            })

    # Calculer les élus au niveau SIRET en utilisant les votes agrégés
    for siret, data in siret_aggregated.items():
        # Calculer le nombre de sièges au niveau SIRET en fonction de l'effectif total
        effectif = data.get("effectif", 0)
        nb_sieges = calculer_nombre_elus_cse(effectif) if effectif > 0 else 0

        # Plafonner à 35 sièges si nécessaire
        if nb_sieges > 35:
            nb_sieges = 35

        data["nb_sieges_cse"] = nb_sieges

        # Récupérer les voix agrégées
        voix_siret = {orga: int(v) for orga, v in data["voix_par_orga"].items() if v > 0}

        if voix_siret and nb_sieges > 0:
            elus_recalcules = repartir_sieges_quotient_seul(voix_siret, nb_sieges)
            data["elus_par_orga"] = defaultdict(int, elus_recalcules)
        else:
            data["elus_par_orga"] = defaultdict(int)

    # Formater les données agrégées pour l'export Excel
    elections_list = []
    for siret_data in siret_aggregated.values():
        # Calculer la participation au niveau SIRET à partir des totaux agrégés
        participation_siret = None
        if siret_data["inscrits"] > 0 and siret_data["votants"] > 0:
            participation_siret = (siret_data["votants"] / siret_data["inscrits"]) * 100

        # Convertir voix_par_orga en all_orgs pour l'affichage
        sve_total = siret_data["sve"]
        all_orgs = []
        for orga, voix in siret_data["voix_par_orga"].items():
            if voix > 0:
                percent = (voix / sve_total * 100) if sve_total > 0 else None
                all_orgs.append({
                    "label": orga,
                    "votes": voix,
                    "percent": percent,
                })

        elections_list.append({
            "siret": siret_data["siret"],
            "raison_sociale": siret_data["raison_sociale"],
            "ud": siret_data["ud"],
            "region": siret_data["region"],
            "effectif": siret_data["effectif_siret"] if siret_data["effectif_siret"] > 0 else None,
            "cycle": siret_data["cycle"],
            "date": siret_data["date"],
            "date_pv": siret_data["date_pv"],
            "institution": siret_data["institution"],
            "fd": siret_data["fd"],
            "idcc": siret_data["idcc"],
            "sve": siret_data["sve"],
            "participation": participation_siret,
            "nb_college": siret_data["nb_college"],
            "all_orgs": all_orgs,
            "nb_sieges_cse": siret_data["nb_sieges_cse"] if siret_data["nb_sieges_cse"] > 0 else None,
            "elus_par_orga": dict(siret_data["elus_par_orga"]),
        })

    # Trier par date
    elections_list = sorted(elections_list, key=lambda x: x["date"])

    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendrier Elections"

    # En-têtes avec style
    headers = [
        "SIRET",
        "Raison sociale",
        "UD",
        "Région",
        "Effectif",
        "Cycle",
        "Date élection",
        "Date PV",
        "Institution",
        "FD",
        "IDCC",
        "SVE",
        "Nb Collèges",
        "Participation (%)",
        # Toutes les organisations (voix + %)
        "CGT - Voix",
        "CGT - %",
        "CFDT - Voix",
        "CFDT - %",
        "FO - Voix",
        "FO - %",
        "CFTC - Voix",
        "CFTC - %",
        "CFE-CGC - Voix",
        "CFE-CGC - %",
        "UNSA - Voix",
        "UNSA - %",
        "Solidaires - Voix",
        "Solidaires - %",
        "Autre - Voix",
        "Autre - %",
        # Élus CSE (moyenne haute - liste complète)
        "Nb sièges CSE (moy. haute)",
        "CGT - Élus (moy. haute)",
        "CFDT - Élus (moy. haute)",
        "FO - Élus (moy. haute)",
        "CFTC - Élus (moy. haute)",
        "CFE-CGC - Élus (moy. haute)",
        "UNSA - Élus (moy. haute)",
        "Solidaires - Élus (moy. haute)",
        "Autre - Élus (moy. haute)",
    ]

    # Note d'avertissement en haut de la feuille
    warning_cell = ws.cell(row=1, column=1, value="⚠️ MOYENNE HAUTE (max 35 sièges) : Les élus CSE sont calculés en supposant que chaque organisation a présenté une liste complète (autant de candidats que de sièges à pourvoir). Plafonné à 35 sièges maximum pour les collèges de 10 000+ inscrits. Le nombre réel d'élus peut être inférieur.")
    warning_cell.font = Font(bold=True, color="FF6B35", size=11)
    warning_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    warning_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells('A1:AM1')  # Fusionner sur toutes les colonnes
    ws.row_dimensions[1].height = 40

    # Style des en-têtes
    header_fill = PatternFill(start_color="D5001C", end_color="D5001C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 15  # SIRET
    ws.column_dimensions['B'].width = 40  # Raison sociale
    ws.column_dimensions['C'].width = 12  # UD
    ws.column_dimensions['D'].width = 20  # Région
    ws.column_dimensions['E'].width = 12  # Effectif
    ws.column_dimensions['F'].width = 10  # Cycle
    ws.column_dimensions['G'].width = 12  # Date élection
    ws.column_dimensions['H'].width = 12  # Date PV
    ws.column_dimensions['I'].width = 12  # Institution
    ws.column_dimensions['J'].width = 10  # FD
    ws.column_dimensions['K'].width = 10  # IDCC
    ws.column_dimensions['L'].width = 12  # SVE
    ws.column_dimensions['M'].width = 12  # Nb Collèges
    ws.column_dimensions['N'].width = 15  # Participation
    # Organisations (8 x 2 colonnes)
    ws.column_dimensions['O'].width = 12  # CGT Voix
    ws.column_dimensions['P'].width = 10  # CGT %
    ws.column_dimensions['Q'].width = 12  # CFDT Voix
    ws.column_dimensions['R'].width = 10  # CFDT %
    ws.column_dimensions['S'].width = 12  # FO Voix
    ws.column_dimensions['T'].width = 10  # FO %
    ws.column_dimensions['U'].width = 12  # CFTC Voix
    ws.column_dimensions['V'].width = 10  # CFTC %
    ws.column_dimensions['W'].width = 12  # CFE-CGC Voix
    ws.column_dimensions['X'].width = 10  # CFE-CGC %
    ws.column_dimensions['Y'].width = 12  # UNSA Voix
    ws.column_dimensions['Z'].width = 10  # UNSA %
    ws.column_dimensions['AA'].width = 13  # Solidaires Voix
    ws.column_dimensions['AB'].width = 10  # Solidaires %
    ws.column_dimensions['AC'].width = 12  # Autre Voix
    ws.column_dimensions['AD'].width = 10  # Autre %
    # Élus CSE
    ws.column_dimensions['AE'].width = 15  # Nb sièges CSE
    ws.column_dimensions['AF'].width = 12  # CGT Élus
    ws.column_dimensions['AG'].width = 12  # CFDT Élus
    ws.column_dimensions['AH'].width = 12  # FO Élus
    ws.column_dimensions['AI'].width = 12  # CFTC Élus
    ws.column_dimensions['AJ'].width = 12  # CFE-CGC Élus
    ws.column_dimensions['AK'].width = 12  # UNSA Élus
    ws.column_dimensions['AL'].width = 13  # Solidaires Élus
    ws.column_dimensions['AM'].width = 12  # Autre Élus

    # Remplir les données (commence à la ligne 3, car ligne 1 = avertissement, ligne 2 = en-têtes)
    for row_num, election in enumerate(elections_list, 3):
        ws.cell(row=row_num, column=1, value=election["siret"])
        ws.cell(row=row_num, column=2, value=election["raison_sociale"])
        ws.cell(row=row_num, column=3, value=election["ud"])
        ws.cell(row=row_num, column=4, value=election["region"])
        ws.cell(row=row_num, column=5, value=election["effectif"])
        ws.cell(row=row_num, column=6, value=election["cycle"])
        ws.cell(row=row_num, column=7, value=election["date"].strftime("%d/%m/%Y") if election["date"] else "")
        ws.cell(row=row_num, column=8, value=election["date_pv"].strftime("%d/%m/%Y") if election["date_pv"] else "")
        ws.cell(row=row_num, column=9, value=election["institution"])
        ws.cell(row=row_num, column=10, value=election["fd"])
        ws.cell(row=row_num, column=11, value=election["idcc"])
        ws.cell(row=row_num, column=12, value=int(election["sve"]) if election["sve"] else None)
        ws.cell(row=row_num, column=13, value=int(election["nb_college"]) if election["nb_college"] else None)
        ws.cell(row=row_num, column=14, value=round(election["participation"], 1) if election["participation"] else None)

        # Toutes les organisations (8 x 2 colonnes)
        all_orgs = election.get("all_orgs", [])
        # Créer un dictionnaire pour accès rapide par label
        orgs_dict = {org["label"]: org for org in all_orgs}

        # CGT (colonnes 15-16)
        cgt = orgs_dict.get("CGT", {})
        ws.cell(row=row_num, column=15, value=int(cgt["votes"]) if cgt.get("votes") else None)
        ws.cell(row=row_num, column=16, value=round(cgt["percent"], 1) if cgt.get("percent") else None)

        # CFDT (colonnes 17-18)
        cfdt = orgs_dict.get("CFDT", {})
        ws.cell(row=row_num, column=17, value=int(cfdt["votes"]) if cfdt.get("votes") else None)
        ws.cell(row=row_num, column=18, value=round(cfdt["percent"], 1) if cfdt.get("percent") else None)

        # FO (colonnes 19-20)
        fo = orgs_dict.get("FO", {})
        ws.cell(row=row_num, column=19, value=int(fo["votes"]) if fo.get("votes") else None)
        ws.cell(row=row_num, column=20, value=round(fo["percent"], 1) if fo.get("percent") else None)

        # CFTC (colonnes 21-22)
        cftc = orgs_dict.get("CFTC", {})
        ws.cell(row=row_num, column=21, value=int(cftc["votes"]) if cftc.get("votes") else None)
        ws.cell(row=row_num, column=22, value=round(cftc["percent"], 1) if cftc.get("percent") else None)

        # CFE-CGC (colonnes 23-24)
        cfe = orgs_dict.get("CFE-CGC", {})
        ws.cell(row=row_num, column=23, value=int(cfe["votes"]) if cfe.get("votes") else None)
        ws.cell(row=row_num, column=24, value=round(cfe["percent"], 1) if cfe.get("percent") else None)

        # UNSA (colonnes 25-26)
        unsa = orgs_dict.get("UNSA", {})
        ws.cell(row=row_num, column=25, value=int(unsa["votes"]) if unsa.get("votes") else None)
        ws.cell(row=row_num, column=26, value=round(unsa["percent"], 1) if unsa.get("percent") else None)

        # Solidaires (colonnes 27-28)
        solidaires = orgs_dict.get("Solidaires", {})
        ws.cell(row=row_num, column=27, value=int(solidaires["votes"]) if solidaires.get("votes") else None)
        ws.cell(row=row_num, column=28, value=round(solidaires["percent"], 1) if solidaires.get("percent") else None)

        # Autre (colonnes 29-30)
        autre = orgs_dict.get("Autre", {})
        ws.cell(row=row_num, column=29, value=int(autre["votes"]) if autre.get("votes") else None)
        ws.cell(row=row_num, column=30, value=round(autre["percent"], 1) if autre.get("percent") else None)

        # Nombre d'élus CSE par organisation (colonnes 31-39)
        ws.cell(row=row_num, column=31, value=election.get("nb_sieges_cse"))

        elus_par_orga = election.get("elus_par_orga", {})
        ws.cell(row=row_num, column=32, value=elus_par_orga.get("CGT"))
        ws.cell(row=row_num, column=33, value=elus_par_orga.get("CFDT"))
        ws.cell(row=row_num, column=34, value=elus_par_orga.get("FO"))
        ws.cell(row=row_num, column=35, value=elus_par_orga.get("CFTC"))
        ws.cell(row=row_num, column=36, value=elus_par_orga.get("CFE-CGC"))
        ws.cell(row=row_num, column=37, value=elus_par_orga.get("UNSA"))
        ws.cell(row=row_num, column=38, value=elus_par_orga.get("Solidaires"))
        ws.cell(row=row_num, column=39, value=elus_par_orga.get("Autre"))

    # Geler les 2 premières lignes (avertissement + en-têtes)
    ws.freeze_panes = "A3"

    # Sauvegarder dans un buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer
